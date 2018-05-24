import os

import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side, NamedStyle

'''
@author Arjun Gupta
@date 5/23/2018
@version 1.0.1

This program takes in two Allegro Netlists that have been converted to Excel format (copy and paste from .htm)
and parses them to accurately and rapidly find all points at both the net and the pin level where there are 
disagreements between the two netlists. The main purpose of this is to aide in the translation of schematics
from one format (in my case, OrCAD) to another (Cadence HDL). 

To do: 
    -Go straight from htm to Excel upload (aka supporting more file formats)
        -Goal of this is to streamline process - instead of having to copy & paste from htm file, you should just 
        be able to drop in the htm files to the Data subfolder.
    
    -Redesign with classes in mind - limit the amount of information reuse
        -Goal of this is to increase overall code readability and decrease runtime for large input files
    
    -Decide on one format - either Pandas or OpenPyxl
        -Goal of this is to reduce the amount of dependencies and O(N^2) passes through the list.
    
Overall efficiency is O(N^2) but true efficiency is θ(12N^2) or higher where N is the amount of rows in the two 
netlists.


Changelist:
Version 1.0.1:
    -Added finishing output statements to show that the program successfully uploaded the new sheets.
    -Added more detail to what is different in Compared Net Results

Version 1.0:
    -Original commit
'''


def main():

    file_path, data, xl = find_file()

    # TODO: scrap pandas and use openpyxl instead to keep one consistent data structure
    sheet1_data = data_frame_to_dict(data[0])
    sheet2_data = data_frame_to_dict(data[1])

    # Grabs two separate sets of info: one dict containing inconsistencies with the pins between the same net
    # and the net names themselves
    diff_dict_pins, diff_dict_nets = compare_sheets(sheet1_data, sheet2_data)

    # Only need to load the workbook once to prevent overwriting data.
    wb = openpyxl.load_workbook(filename=file_path)

    # NamedStyles belong to one workbook - no need to create the same NamedStyle twice
    if "BorderAndFont" not in wb.named_styles:
        wb.add_named_style(create_font_style())

    sheet1_name = "Compared Pin Results"
    sheet2_name = "Compared Net Results"

    #
    export(file_path, diff_dict_pins, sheet1_name, wb)
    export(file_path, diff_dict_nets, sheet2_name, wb)

    print("Export of \"{}\" and \"{}\" to \"{}\" is successful!".format(sheet1_name, sheet2_name, file_path))


# Ensures that the user is only able to input valid files/sheets
def find_file():
    subdir = "Data"
    path = os.path.normpath(os.path.join(os.getcwd(), subdir))
    files = os.listdir(path)
    print("Current files: ")
    print("-------------------------------------------------------------")
    for f in files:
        # TODO: Create a file selection prompt, selecting from numbers rather than having to enter the file name
        print(f)
    print("-------------------------------------------------------------")
    # Need a try/catch here for above to do
    file = input("Enter the file you want to read from: ")
    data_frames = None
    prev_path = path
    # Excel file
    xl = None
    # Validation of proper file entered
    while data_frames is None:
        try:
            path = os.path.normpath(os.path.join(path, file))
            extension = file.split(".")
            # Ex: User enters a directory rather than a file
            if len(extension) > 1:
                if extension[1] == "xls" or extension[1] == "xlsx":
                    # convert xlsx to csv here
                    xl, data_frames = excel_to_dataframe(path)
                # TODO: Add support for htm
                else:
                    raise IOError("Unsupported file type selected.")
            else:
                raise IOError("Improper file name entered.")
        # Raise if unable to find the file in the subfolder
        except OSError:
            print('Unable to find file.')
            file = input("Enter the file name that you want to read from: ")
            path = prev_path
        except IOError:
            file = input("Enter the file name that you want to read from: ")
            path = prev_path
    # Return the path, the populated dataframe, and the workbook
    return path, data_frames, xl


# Parses the excel file and converts it into a Pandas DataFrame
def excel_to_dataframe(path):
    # Path verified in above file conditioning
    xl = pd.ExcelFile(path)
    sheet_names = xl.sheet_names

    # Prevents passing in a excel book with only one sheet.
    if len(sheet_names) <= 1:
        raise IOError("Improper file format.")

    print("Sheets available for parsing: ", *sheet_names, sep='\t')
    sheet1 = input("Enter first sheet to parse: ")
    sheet2 = input("Enter second sheet to parse: ")
    valid_sheets = False

    # Ensures sheets are valid since Pandas requires a string representing the sheet name
    while not valid_sheets:

        if sheet1 not in sheet_names and sheet2 not in sheet_names:
            print("Both sheets were not found.\n")
            sheet1 = input("Enter first sheet to parse: ")
            sheet2 = input("Enter second to parse: ")
            continue

        elif sheet1 not in sheet_names:
            print("The first sheet was not found.\n")
            sheet1 = input("Enter first sheet to parse, second sheet is \'{}\': ".format(sheet2))
            continue

        elif sheet2 not in sheet_names:
            print("The second sheet was not found.\n")
            sheet2 = input("Enter second sheet to parse, first sheet is \'{}\': ".format(sheet1))
            continue

        elif sheet1 == sheet2:
            print("Duplicate sheet entered!\n")
            sheet2 = input("Enter a new sheet to parse, first sheet is \'{}\': ".format(sheet1))
            continue
        else:
            valid_sheets = True

    print("Sheets being compared are: {} and {}".format(sheet1, sheet2))

    sheet1_parsed = xl.parse(sheet1)
    cols = ["Net Name", "Net Pins"]
    scols = set(cols)
    svals1 = set(sheet1_parsed.columns.values.tolist())
    # Sometimes there will be extra text at the top of the sheet that isn't "Net Name" or "Net Pins"
    # we need to delete this if its not
    if not scols.issubset(svals1):
        sheet1_parsed = sheet1_parsed.drop([0])
        sheet1_parsed.columns = cols

    sheet2_parsed = xl.parse(sheet2)
    svals2 = set(sheet2_parsed.columns.values.tolist())
    if not scols.issubset(svals2):
        sheet2_parsed = sheet2_parsed.drop([0])
        sheet2_parsed.columns = cols

    return xl, [sheet1_parsed, sheet2_parsed]


# Converts Pandas dataframe to Dictionary with Key, Value format as follows: {Key -> String, Value -> [List-of String]}
def data_frame_to_dict(frame):
    frame_dict = {}
    for row in frame.values:
        if len(row) == 2:
            net_name = row[0]
            # Need to split so that each individual pin can be compared
            split_ref_des = row[1].split(" ")
            frame_dict[net_name] = split_ref_des
    return frame_dict


# Compares the two sheets and outputs two different dictionaries representing inconsistencies between
# the net pins and net names
def compare_sheets(sheet1, sheet2):
    if sheet1 is None or sheet2 is None:
        raise IOError("Unable to parse due to an error.")

    if (not isinstance(sheet1, dict)) or (not isinstance(sheet2, dict)):
        raise IOError("Unable to parse something that is not a dictionary")

    diff_dict_pins = {}
    diff_dict_nets = {}
    for net_name1, net_name2 in zip(sheet1, sheet2):
        if net_name1 in sheet2:
            # Create a tuple of (Sheet1CellValue, Sheet2CellValue)
            # This is used to easily show the differences between the two sheets
            for val1, val2 in zip(sheet1[net_name1], sheet2[net_name1]):
                if val1 not in sheet2[net_name1]:
                    if net_name1 not in diff_dict_pins:
                        diff_dict_pins[net_name1] = []
                    diff_dict_pins[net_name1].append("[ {} v {} ]".format(val1, val2))
        # Triggers if a net name in the first compared sheet is not in the second compared sheet
        else:
            # change to make more elegant...
            diff_dict_nets["{} not in second sheet".format(net_name1)] = sheet1[net_name1]
            if net_name2 not in sheet1:
                diff_dict_nets["{} not in first sheet".format(net_name2)] = sheet2[net_name2]

            # print("Key: {} \t Value: {}".format(net_name, diff_dict_nets[net_name]))
    return diff_dict_pins, diff_dict_nets


# Exports the dictionary to the Excel Spreadsheet under the given sheet name
def export(file_path, diff_dict, sheet_name, wb):
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
        print(wb.sheetnames)
    ws = wb[sheet_name]

    # Bolds the font for the "Net Name" and "Net Pins" Column (Col A1 and B1 respectively)
    font_column = Font(name='Times New Roman',
                       size=10,
                       bold=True)

    ws['A1'] = "Net Name"
    ws['B1'] = "Net Pins"
    ws['A1'].style = "BorderAndFont"
    ws['B1'].style = "BorderAndFont"
    ws['A1'].font = font_column
    ws['B1'].font = font_column
    row = 1
    col = 1

    # Assigns the specific net name and pins matching to the corresponding cell
    for key in diff_dict.keys():
        row += 1
        # assigning the specific cell that is going be mutated
        net_name = ws.cell(row=row, column=col)
        net_pins = ws.cell(row=row, column=col+1)
        net_name.value = key
        # Can only have hashable values (Strings, integers, etc., NOT lists) as values in a cell
        net_pins.value = " ".join(diff_dict[key])

        net_name.style = "BorderAndFont"
        net_pins.style = "BorderAndFont"

        # print("Key: {} \t Value: {}".format(key, diff_dict[key]))

    # file_name = ""
    # for line in file_path:
    #     drive, path = os.path.splitdrive(line)
    #     path, file_name = os.path.split(path)

    # Trying to make  the column width and row height easy to view
    # Need to update...
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                width = max((dims.get(cell.column, 0), len(cell.value)))
                # Padding row dimensions, look into executing a vba script
                if width >= 125:
                    factor = width // 125
                    ws.row_dimensions[cell.row].height = 15 * factor
                    continue
                dims[cell.column] = width
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # NOTE: Program crashes when file that is being saved is open... Don't do that please
    wb.save(filename=file_path)
    wb.close()

    # TODO: Look into a solution for this? PermissionError raised but I don't know where
    # try:
    #     wb.save(filename="temp.xlsx")
    # except PermissionError:
    #     print("HERE")
    #     print(os.access("temp.xlsx", os.W_OK))
    #     while not os.access(file_path, (os.F_OK ^ os.R_OK ^ os.W_OK)):
    #         input("Waiting for user to close {}. Press Enter once done!".format(file_name))


def create_font_style():
    border_style = NamedStyle(name="BorderAndFont")
    border_style.font = Font(name='Times New Roman',
                             size=10,
                             color='000000')
    bd = Side(style='thin', color='000000')
    border_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    border_style.alignment.wrap_text = True
    return border_style


if __name__ == '__main__':
    main()
